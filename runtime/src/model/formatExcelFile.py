import openpyxl as xl
import datetime

class ExcelReader:
    
    def __init__(self,
                 filepath: str = 'schedule.xlsx'):
        self.filepath = filepath
        self.key = {}

    def getKey(self):
        return self.key
    
    def setFilepath(self, filepath):
        self.filepath = filepath

    def getTodaysSchedule(self):
        workbook = xl.load_workbook(self.filepath)
        sheet = workbook.active
        merged_cells = sheet.merged_cells
        mergedCellsList = []

        for i in merged_cells:
            mergedCellsList.append(str(i))

        self.key = {}
        noKey = True
        recordKey = False
        keyOnRow = 0

        # recording the key (Must be at the top of the excel file)
        # ASSURE THAT ALL CELL COLORS ARE FORMATTED CORRECTLY OR THE PROGRAM WILL JUST DIE
        for rowNum, row in enumerate(sheet.iter_rows()):
            for cell in row:
                if noKey:
                    if cell.value == 'Key':
                        recordKey = True
                        keyOnRow = rowNum
                        noKey = False
                        continue
                elif recordKey:
                    if cell.fill:
                        self.key[str(cell.fill.start_color.rgb)] = cell.value
            recordKey = False
            if not noKey and not recordKey:
                break

        breakOut = False            # Have we found our week and fully worked through it, and can skip the rest of the sheet?
        skipRow = False             # Does the row need skipped?
        weekFound = False           # has this week been found?

        # Find today, and find Monday of that given week
        # Thanks Google Bard
        today = datetime.date.today()
        beginOfWeek = today - datetime.timedelta(days=today.weekday())
        theWeek = []

        # find only this weeks schedule
        for rowNum, row in enumerate(sheet.iter_rows()):
            weekRow = []
            cellToMerge = None
            merge = False
            if rowNum <= keyOnRow:
                continue
            for cell in row:
                if not skipRow:
                    if cell.value == None and cell.column == 1:         # Specific case for if the first column is empty (at the top of the excel sheet there's a blank space
                        skipRow = True                                  #                                                 next to the day of the week)
                        continue
                    elif cell.column == 1 and not weekFound:            # Skip the month column entry
                        continue 
                    elif cell.column == 2 and str(cell.value).split(' ')[0] == str(beginOfWeek):  # Look for a date, if it is the beginning of the week, we found what we need
                        weekFound = True                                # so we set weekFound
                    if cell.value == 'NOTES' and weekFound:             # NOTES is at the end of the week, so we know we can breakout
                        breakOut = True
                    if weekFound:                                       # We have found our week, so this is the data we care about
                        if not merge:                                   # If we don't have a merged cell, this cell should exist in our data
                            weekRow.append(cell)
                        elif merge:                                     # If we do have a merged cell, we need to populate it through the rest of the cols which would otherwise be 'None'
                            weekRow.append(cellToMerge)
                        for string in mergedCellsList:
                            cells = string.split(':')                   # Split the possible mergedCells into their starting and ending coords
                            if cell.coordinate in cells:                # If our current cells coordinates are in that list, we know we have a merged cell
                                if not merge:                   
                                    merge = True                        # set the merge bit to True, to tell the loop to populate the merge cell later
                                    cellToMerge = cell
                                elif merge:
                                    merge = False                       # Once we find the end of our merged cell, set the bit back to False so we don't populate it further
                                    cellToMerge = None
                            else:
                                pass
            
            if weekFound:                                   # If we found our week,         
                theWeek.append(weekRow)                     # add each row to theWeek
            skipRow = False
            if breakOut:
                breakOut = False
                break

        dateList = []                                       # Find and decode the dates of the week
        for i in theWeek[0]:
            dateList.append(str(i.value).split(' ')[0])

        todaysIndex = dateList.index(str(today)) + 1        # Find today's index in the list

        today = []

        for rowNum, row in enumerate(theWeek):              # Use today's index to find the schedule for today
            if rowNum == 0:
                today.append(str(row[todaysIndex - 1].value).split(' ')[0])
            else:
                entry = []                                  # format an entry as a list
                entry.append(row[0].value)
                entry.append(row[todaysIndex].value)        # add each datapoint to it
                if row[todaysIndex].fill.start_color.rgb in self.key:
                    entry.append(str(self.key[str(row[todaysIndex].fill.start_color.rgb)]))
                else:
                    entry.append('No Location')
                today.append(entry)                         # add the entry to today

        return today                                        # return today's schedule
    